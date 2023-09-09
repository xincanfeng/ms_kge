import openpyxl
from openpyxl.styles import Font, colors, Alignment, PatternFill
import os
import sys
import heapq
import pandas as pd

# import torch
# print(torch.__version__, torch.version.cuda)


# directory:
# |-- MS-KGE
#  ------ make_table.py
#  ------ exp_sampling.xlsx
#  ------ models
#    | ------- model_dataset_id
#      | ---------- train.log


datasets = ['FB15k-237', 'wn18rr', 'FB15k', 'wn18', 'YAGO3-10']
models = ['HAKE', 'ComplEx', 'RotatE', 'pRotatE', 'TransE', 'DistMult']
# copy the row of models manually
model_rows = [3, 52, 119, 186, 193, 200, 207]


if len(sys.argv) > 1:
    var_name = sys.argv[1]
else:
    var_name = 'now'


wb = openpyxl.load_workbook(filename='results/temp.xlsx')
# default sheet is the first sheet
ws1 = wb.active
ws2 = wb.copy_worksheet(ws1)
wb.move_sheet(ws2, -4)


def _eval_target_lines_and_ws(eval_target):
    if eval_target == 'Test':
        ws = ws1
        ws.title = 'eval_test'
        ws.sheet_properties.tabColor = "76D6FF"
        ws.sheet_view.zoomScale = 110
    if eval_target == 'Valid':
        ws = ws2
        ws.title = 'eval_valid'
        ws.sheet_properties.tabColor = "7A81FF"
        ws.sheet_view.zoomScale = 110
    return eval_target, ws


# read list_script from excel file.
# dict_script is a dict
# (key: filename, value: bash_script)
# read the Test/Valid score
# trim the original file
# Test/Valid line -5, -4, -3, -2, -1
eval_lines = (5, 0, -1)
def _read_txt_rstrip_keep_first_and_eval_lines(url, eval_target):
    file = open(url, 'r', encoding='utf-8')
    eval_target = _eval_target_lines_and_ws(eval_target)[0]
    lines = file.readlines()
    file.close()
    seg_list = []
    # append the first_line
    seg_list.append(lines[0].rstrip(" \n"))
    # append the eval_lines
    lines2 = []
    # print(len(lines), eval_target, url)
    for j in range(len(lines)):
        # if train.log is already processed to eval_target
        if eval_target in lines[j] and lines[j].rstrip(" \n").split(" ")[-1]!='Dataset...':
            lines2.append(lines[j])
    # skip if data is not prepared
    if len(lines2) >= 5:
        for i in range(*eval_lines):
            seg_list.append(lines2[-i].rstrip(" \n"))
            pass
    # seg_list.reverse()
    return seg_list


# read data
# initialize a dict to save data
# data: {key: filename, value: [MRR, MR, HITS@1, HITS@3, HITS@10]}
# slice [1:6] corresponding to temperature [2:0.01]
def _read_file_values_into_dic(eval_target):
    data = {}
    eval_target = _eval_target_lines_and_ws(eval_target)[0]
    for model_index, model in enumerate(models):
        for dataset_index, dataset in enumerate(datasets):
            # compute the start point of column and row
            row_start = model_rows[model_index]
            row_end = model_rows[model_index+1]
            id_exp = -1
            for row_exp in range (row_start, row_end):
                id_exp = id_exp + 1 
                row_orientation = row_start + id_exp
                column_orientation = 9 + 8 * dataset_index
                filename = model + '_' + dataset + '_' + str(id_exp)
                filepath = os.path.join('models', filename, 'train.log')
                if os.path.exists(filepath):
                    row_data = _read_txt_rstrip_keep_first_and_eval_lines(filepath, eval_target)
                    # tmp_list by saving sequence is [start_time, end_time, MRR, MR, HITS@1, HITS@3, HITS@10]
                    tmp_list = []
                    # append the train.log start time
                    tmp_start = row_data[0].split(" ")
                    tmp_list.append('{}_{}'.format(tmp_start[0], tmp_start[1]))
                    # if not, the tmp_list has no elements
                    # when data is not prepared, jump
                    if len(row_data) > 1:
                        if eval_target in row_data[1]:
                            # append the train.log end time
                            tmp_end = row_data[1].split(" ")
                            tmp_list.append('{}_{}'.format(tmp_end[0], tmp_end[1]))
                            # append MRR, MR, HITS@N
                            for line in row_data[1:]:
                                value = float(line.split(" ")[-1])
                                tmp_list.append(value)
                    data[filename] = tmp_list
                else:
                    data[filename] = None
    return data
            

mrr_column = [11, 19, 27, 35, 43]
mr_column = [i+1 for i in mrr_column]
hit_one_column = [i+2 for i in mrr_column]
hit_three_column = [i+3 for i in mrr_column]
hit_ten_column = [i+4 for i in mrr_column]


# larger value of MRR, H@N means better
# define the max_num needed manually, e.g., max_num=3
def _add_font_to_max(metric_column, max_num, ws):
    for j in metric_column:
        # find out the values we want to highlight and put them in a list (assume each value is unique) 
        value_list = []
        # find base on some model
        for i in range(model_rows[0], model_rows[6]):
            # read value from excel, instead of train.log
            column_value = ws.cell(row=i, column=j).value
            # skip if there is missed values
            if column_value is None: continue
            value_list.append(column_value)
        # select the two largest values in the results of the same model 
        column_dmax = heapq.nlargest(2*max_num, value_list)
        # select the largest value in the results of the same model 
        column_max = heapq.nlargest(max_num, value_list)
        for k in range(model_rows[0], model_rows[6]):
            if ws.cell(row=k, column=j).value in column_dmax:
                # make bold
                ws.cell(row=k, column=j).font = Font('Arial', bold=True, size=12)
                if ws.cell(row=k, column=j).value in column_max:
                    # make bold and color
                    ws.cell(row=k, column=j).font = Font('Arial', bold=True, color='0433FF', size=12)


# smaller value in MR means better
def _add_font_to_min(metric_column, min_num, ws):
    for j in metric_column:
        value_list = []
        for i in range(model_rows[0], model_rows[6]):
            column_value = ws.cell(row=i, column=j).value
            # skip
            if column_value is None: continue
            if column_value == '/': continue
            value_list.append(column_value)
        column_dmin = heapq.nsmallest(2*min_num, value_list)
        column_min = heapq.nsmallest(min_num, value_list)
        for k in range(model_rows[0], model_rows[6]):
            if ws.cell(row=k, column=j).value in column_dmin:
                ws.cell(row=k, column=j).font = Font('Arial', bold=True, size=12)
                if ws.cell(row=k, column=j).value in column_min:
                    ws.cell(row=k, column=j).font = Font('Arial', bold=True, color='0433FF', size=12)


# experimental_set_column = [6, 5, 4, 3]
def _add_style_to_max_in_smallest_experimental_set(metric_column, ws):
    # find the smallest_experimental_set_num
    # s: the experimental_set starting column
    s = 6
    for t in range(model_rows[0], model_rows[6]):
        if t in model_rows: continue
        # find set_column
        set_value = ws.cell(row=t, column=s).value
        if set_value == '-':
            set_column = s - 1
        else:
            set_column = s
        # find set_lnext_column
        set_lnext_value = ws.cell(row=t, column=set_column-1).value
        if set_lnext_value is None: continue
        elif set_lnext_value == '-':
            if ws.cell(row=t, column=set_column-2).value is None: continue
            set_lnext_column = set_column - 2
        else:
            set_lnext_column = set_column - 1
        # find set_lnext_num
        for c in range(7):
            set_lnext_dnext_value = ws.cell(row=t+c+1, column=set_lnext_column).value
            if set_lnext_dnext_value is None:
                # c = 0, c+1=1, c+1+1=2
                c = c + 1
            else: break
        set_lnext_num = c + 1
        # add style to max in the smallest experimental set acording to the metric_column
        for j in metric_column:
            value_list = []
            # e.g., t=1, t+3=4
            # e.g., i=1, 2, 3
            for i in range(t, t+set_lnext_num):
                column_value = ws.cell(row=i, column=j).value
                # skip
                if column_value is None: continue
                value_list.append(column_value)
            max_in_set = max(value_list, default=0)
            for k in range(t, t+set_lnext_num):
                if ws.cell(row=k, column=j).value == max_in_set:
                    # add PatternFill
                    ws.cell(row=k, column=j).fill = PatternFill('solid', start_color='73FDD6')


# fill in the excel
def fill_in_the_excel(eval_target):
    data = _read_file_values_into_dic(eval_target)
    ws = _eval_target_lines_and_ws(eval_target)[1]
    for model_index, model in enumerate(models):
        for dataset_index, dataset in enumerate(datasets):
            # compute the start point of column and row
            row_start = model_rows[model_index]
            row_end = model_rows[model_index+1]
            id_exp = -1
            for row in range (row_start, row_end):
                id_exp = id_exp + 1
                row_orientation = row_start + id_exp
                # the first columns are: 8, 16, 24, ...
                column_orientation = 9 + 8 * dataset_index
                filename = model + '_' + dataset + '_' + str(id_exp)
                # if train.log doesn't exist
                if data[filename] is None:
                    ws.cell(row=row_orientation, column=column_orientation).value = 'No file'
                    ws.cell(row=row_orientation, column=column_orientation).font = Font(color='919191')
                    ws.cell(row=row_orientation, column=column_orientation).alignment = Alignment(wrap_text=True)
                # if train.log only contains one record
                elif len(data[filename]) == 1:
                    # fill in the start_time
                    ws.cell(row=row_orientation, column=column_orientation).value = data[filename][0]
                    ws.cell(row=row_orientation, column=column_orientation).font = Font(color='919191')
                    ws.cell(row=row_orientation, column=column_orientation).alignment = Alignment(horizontal='right')
                    # fill in the msg
                    ws.cell(row=row_orientation, column=column_orientation + 1).value = 'Check train.log'
                    ws.cell(row=row_orientation, column=column_orientation + 1).font = Font(color='919191')
                    ws.cell(row=row_orientation, column=column_orientation + 1).alignment = Alignment(wrap_text=True)
                # if train.log contains more than one record
                else:
                    for i in range(7):
                        # fill in the start_time, end_time, MRR, MR ...
                        ws.cell(row=row_orientation, column=(column_orientation + i)).value = data[filename][i]
                        ws.cell(row=row_orientation, column=(column_orientation + i)).alignment = Alignment(horizontal='right')
                        # change the font of the start_time and end_time
                        for i in range(2):
                            ws.cell(row=row_orientation, column=(column_orientation + i)).font = Font(color='919191')
    _add_font_to_max(mrr_column, 3, ws)
    _add_font_to_min(mr_column, 3, ws)
    _add_font_to_max(hit_one_column, 3, ws)
    _add_font_to_max(hit_three_column, 3, ws)
    _add_font_to_max(hit_ten_column, 3, ws)
    _add_style_to_max_in_smallest_experimental_set(mrr_column, ws)
    

# eval_targets = ['Test', 'Valid']
fill_in_the_excel('Test')
fill_in_the_excel('Valid')


# command: python3 make_table.py {$var_name}
wb.save('results/{}_exp_subsampling.xlsx'.format(var_name))